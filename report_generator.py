import re
import difflib
import shutil
import subprocess
import tempfile
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

PRIMARY_COLS = [get_column_letter(c) for c in range(3, 19)]   # C..R (16)
SECONDARY_COLS = [get_column_letter(c) for c in range(20, 28)]  # T..AA (8)

# Kapasitas "default" formulir adalah 16 responden (kolom C..R).
# Jika jumlah responden lebih dari itu, kolom cadangan T..AA (8 kolom) otomatis
# dipakai sehingga total kapasitas template menjadi 24 responden.
DEFAULT_CAPACITY = len(PRIMARY_COLS)                     # 16
MAX_RESPONDENTS = DEFAULT_CAPACITY + len(SECONDARY_COLS)  # 24

SCORE_COL = "S"
KET_COL = "AC"

MASTER_SHEET = "KOLOM ISIAN"

VIEW_SHEETS = ["Materi Pelatihan", "Program Pelatihan", "INSTRUKTUR", "Penyelenggaraan"]

SHEET_TITLE_ROWS = {
    "KOLOM ISIAN": (6, 29),
    "Histogram": (6, 4),
    "Materi Pelatihan": (6, 29),
    "Program Pelatihan": (6, 29),
    "INSTRUKTUR": (7, 29),
    "Penyelenggaraan": (7, 29),
    "NILAI INST": (6, 6),
}

TITLE_TEXT_CELLS = {
    "Materi Pelatihan": "A6",
    "Program Pelatihan": "A6",
    "INSTRUKTUR": "A7",
    "Penyelenggaraan": "A7",
}

REPORT_TITLE = "HASIL ANALISIS EVALUASI"

HEADER_ROWS = [13, 22, 30, 38]

SECTION_LABELS = {
    "Materi Pelatihan": ["Info mudah didapat", "Pendaftaran mudah", "Petunjuk pendaftaran jelas", "Program jelas"],
    "Program Pelatihan": ["Program menarik", "Program bermanfaat", "Kompetensi meningkat", "Durasi sesuai"],
    "INSTRUKTUR": ["Instruktur menguasai materi", "Kemampuan penyampaian instruktur", "Kemampuan mengelola peserta", "Sikap & teladan instruktur"],
    "Penyelenggaraan": ["Pelayanan petugas", "Jadwal sesuai rencana", "Perlengkapan tepat waktu", "Sarana pelatihan memadai", "Sarana penunjang memadai"],
}


def _normalize(text: str) -> str:
    text = str(text).lower().replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _find_question_row(ws, question_text: str, cutoff: float = 0.8):
    """Cari baris di kolom B yang cocok dengan teks pertanyaan (fuzzy)."""
    target = _normalize(question_text)
    best_row, best_ratio = None, 0.0
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if not cell.value:
            continue
        norm = _normalize(cell.value)
        # Baris judul section (mis. "PROGRAM PELATIHAN", "INSTRUKTUR") bukan
        # baris pertanyaan asli -- semua pertanyaan asli di template ini
        # diawali "Apakah" atau "Bagaimana". Lewati baris yang bukan
        # pertanyaan supaya tidak pernah ikut cocok lewat jalan pintas
        # fuzzy-substring di bawah.
        if not (norm.startswith("apakah") or norm.startswith("bagaimana")):
            continue
        if norm == target:
            return cell.row
        ratio = difflib.SequenceMatcher(None, norm, target).ratio()
        if norm in target or target in norm:
            ratio = max(ratio, 0.9)
        if ratio > best_ratio:
            best_ratio, best_row = ratio, cell.row
    return best_row if best_ratio >= cutoff else None


def kategori_nilai(score: float) -> str:
    if score >= 5:
        return "SANGAT BAIK"
    if score >= 4:
        return "BAIK"
    if score >= 3:
        return "CUKUP"
    if score >= 2:
        return "KURANG"
    return "TIDAK BAIK"


COMMENT_RECAP_HEADERS = [
    "NO.",
    "Dari mana anda mendapatkan informasi tentang pelatihan?",
    "Komentar dan saran Anda terhadap Pelatihan ",
    "Komentar dan saran Anda terhadap Instruktur",
    "Komentar dan saran Anda terhadap penyelenggaraan pelatihan ",
    "Keluhan terhadap penyelenggaraan pelatihan ",
]


def generate_comment_recap(
    df_filtered,
    info_col: str | None,
    comment_col_map: dict,
    training_comment_label: str = "Komentar/saran program pelatihan",
    instructor_comment_label: str = "Komentar/saran instruktur",
    organizer_comment_label: str = "Komentar/saran penyelenggaraan",
    complaint_label: str = "Keluhan penyelenggaraan",
) -> bytes:
    """Buat rekap komentar 1-baris-per-responden (TIDAK membuang responden
    yang kolom komentarnya kosong -- jumlah baris harus selalu sama dengan
    jumlah responden pada data terfilter), dengan header & urutan kolom
    persis format resmi:
    NO. | Sumber informasi | Komentar Pelatihan | Komentar Instruktur |
    Komentar Penyelenggaraan | Keluhan Penyelenggaraan.
    """
    n = len(df_filtered)

    def get_col(label):
        col = comment_col_map.get(label)
        if col and col in df_filtered.columns:
            return df_filtered[col].astype(str).where(df_filtered[col].notna(), "").tolist()
        return [""] * n

    info_values = (
        df_filtered[info_col].astype(str).where(df_filtered[info_col].notna(), "").tolist()
        if info_col and info_col in df_filtered.columns
        else [""] * n
    )

    data = {
        COMMENT_RECAP_HEADERS[0]: list(range(1, n + 1)),
        COMMENT_RECAP_HEADERS[1]: info_values,
        COMMENT_RECAP_HEADERS[2]: get_col(training_comment_label),
        COMMENT_RECAP_HEADERS[3]: get_col(instructor_comment_label),
        COMMENT_RECAP_HEADERS[4]: get_col(organizer_comment_label),
        COMMENT_RECAP_HEADERS[5]: get_col(complaint_label),
    }
    import pandas as pd
    df_out = pd.DataFrame(data, columns=COMMENT_RECAP_HEADERS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Komentar"

    from openpyxl.styles import Font
    ws.append(COMMENT_RECAP_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in df_out.itertuples(index=False):
        ws.append(list(row))

    widths = [5.5, 27.5, 27, 31, 32, 31]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    from openpyxl.styles import Alignment
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _format_training_title(training_title: str) -> str:
    name = re.sub(r"\s+", " ", str(training_title).strip()).upper()
    if not name:
        return ""
    name = re.sub(r"^PELATIHAN BERBASIS KOMPETENSI KEJURUAN\s+", "", name)
    name = re.sub(r"^PELATIHAN KOMPETENSI BERBASIS\s+", "", name)  # kompatibel data lama
    return f"PELATIHAN BERBASIS KOMPETENSI KEJURUAN {name}"


def _format_program_title(program_name: str) -> str:
    name = re.sub(r"\s+", " ", str(program_name).strip()).upper()
    if not name:
        return ""
    name = re.sub(r"^PROGRAM PELATIHAN DASAR\s+", "", name)
    name = re.sub(r"\s*\([^)]*\)", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    if "PAKET" not in name:
        match = re.match(r"^(.*\S)\s+(\d+)$", name)
        if match:
            name = f"{match.group(1)} PAKET {match.group(2)}"
    return f"PROGRAM PELATIHAN DASAR {name}"


def _add_title_separator_line(ws, row: int = 6, min_col: int = 1, max_col: int = 29):
    thin_double = Side(style="double", color="000000")
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        cell.border = Border(
            top=thin_double, bottom=existing.bottom,
            left=existing.left, right=existing.right,
        )


def _sync_titles_and_separator_lines(wb):
    for sheet_name, (row, max_col) in SHEET_TITLE_ROWS.items():
        if sheet_name not in wb.sheetnames:
            continue
        _add_title_separator_line(wb[sheet_name], row=row, max_col=max_col)
    for sheet_name, cell_coord in TITLE_TEXT_CELLS.items():
        if sheet_name not in wb.sheetnames:
            continue
        wb[sheet_name][cell_coord] = REPORT_TITLE


def _fix_secondary_header_numbers(ws):
    for row in HEADER_ROWS:
        if ws.cell(row=row, column=19).value != "JUMLAH":
            continue
        for i, col in enumerate(SECONDARY_COLS):
            ws[f"{col}{row}"] = DEFAULT_CAPACITY + i + 1


def _update_divisor_everywhere(wb, n_respondents: int, use_secondary: bool):
    sum_range = f"C{{r}}:R{{r}},T{{r}}:AA{{r}}" if use_secondary else "C{r}:R{r}"
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=19, max_col=19):
            cell = row[0]
            if isinstance(cell.value, str) and cell.value.startswith("=SUM(") and "/12" in cell.value:
                r = cell.row
                cell.value = f"=SUM({sum_range.format(r=r)})/{n_respondents}"


def _style_histogram_chart(wb, training_title: str, section_values=None):
    if "Histogram" not in wb.sheetnames:
        return
    ws = wb["Histogram"]
    if not ws._charts:
        return
    chart = ws._charts[0]
    try:
        runs = chart.title.tx.rich.p[0].r
        label = re.sub(r"\s+", " ", str(training_title).strip()).upper()
        for run in runs:
            if run.t and "NAMA PELATIHAN" in run.t:
                run.t = f" ({label})" if label else " (NAMA PELATIHAN)"
            if run.rPr is not None:
                run.rPr.u = None
                run.rPr.b = True
    except (AttributeError, IndexError):
        pass
    for ser in chart.series:
        ser.graphicalProperties.solidFill = "595959"
        ser.graphicalProperties.line.noFill = True

    # Skala sumbu Y menyesuaikan data aktual (bukan hardcode), supaya batang
    # selalu terlihat walau nilainya di luar rentang 4.3-4.8 (mis. kategori
    # CUKUP/KURANG). Tetap "zoom in" ala contoh referensi, tapi dinamis.
    if section_values:
        lo = min(section_values)
        hi = max(section_values)
    else:
        lo, hi = 0, 5
    import math
    axis_min = max(0, math.floor(lo * 10) / 10 - 0.2)
    axis_max = min(5, math.ceil(hi * 10) / 10 + 0.2)
    if axis_max - axis_min < 0.5:
        axis_max = min(5, axis_min + 0.5)
    chart.y_axis.scaling.min = round(axis_min, 1)
    chart.y_axis.scaling.max = round(axis_max, 1)
    chart.y_axis.majorUnit = 0.1


def convert_xlsx_to_pdf(xlsx_bytes: bytes, timeout: int = 90) -> bytes:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError(
            "LibreOffice ('soffice') tidak ditemukan di server. "
            "Tambahkan 'libreoffice-calc' ke packages.txt agar konversi ke PDF bisa jalan."
        )
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        in_path = tmp_dir_path / f"{uuid.uuid4().hex}.xlsx"
        in_path.write_bytes(xlsx_bytes)
        result = subprocess.run(
            [
                soffice, "--headless", "--norestore",
                "--convert-to", "pdf:calc_pdf_Export",
                "--outdir", str(tmp_dir_path),
                str(in_path),
            ],
            capture_output=True, timeout=timeout,
        )
        out_path = in_path.with_suffix(".pdf")
        if not out_path.exists():
            stderr = result.stderr.decode("utf-8", errors="ignore")
            raise RuntimeError(
                f"Konversi ke PDF gagal. Detail teknis: "
                f"{stderr or result.stdout.decode('utf-8', errors='ignore')}"
            )
        return out_path.read_bytes()


def _rewrite_cross_sheet_formulas(wb, row_map: dict, canonical_questions: dict, used_cols: list):
    """PERBAIKAN BUG UTAMA: openpyxl kadang merusak formula 'shared formula'
    bawaan Excel saat file dibuka lalu disimpan ulang -- rujukan sel tunggal
    (mis. ='KOLOM ISIAN'!O41) bisa berubah jadi rentang yang salah
    (mis. ='KOLOM ISIAN'!O39:AD39), sehingga sheet turunan membaca baris yang
    keliru.

    Solusi: TULIS ULANG secara eksplisit & deterministik setiap formula
    rujukan-silang di sheet turunan (Materi Pelatihan, Program Pelatihan,
    INSTRUKTUR, Penyelenggaraan) untuk kolom-kolom responden, terlepas dari
    apa pun isi formula aslinya -- formula POLOS satu-sel
    (mis. ='KOLOM ISIAN'!C34), BUKAN dibungkus IF() atau apa pun yang bisa
    memicu Excel mengubahnya jadi rujukan rentang aneh (mis. '=@...O25:AD25').

    Tampilan '0' pada sel yang seharusnya kosong (karena rumus merujuk sel
    kosong di KOLOM ISIAN, dan Excel menampilkan sel kosong yang dirujuk
    sebagai 0) diatasi lewat FORMAT ANGKA custom ('0;-0;;@' -- bagian 'nol'
    dikosongkan), bukan lewat logika formula.
    """
    hide_zero_format = "0;-0;;@"
    for sheet_name in VIEW_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # PENTING: hanya proses label yang MEMANG milik sheet ini (lihat
        # SECTION_LABELS). Sebelumnya kode ini mencoba mencocokkan SEMUA 17
        # pertanyaan ke SETIAP sheet turunan -- akibatnya, pertanyaan milik
        # section lain (mis. "Program menarik" dari sheet "Program Pelatihan")
        # ikut dicari di sheet "Penyelenggaraan", dan karena _find_question_row
        # punya jalan pintas fuzzy-substring, ia bisa salah nempel ke baris
        # judul section (mis. baris header "II PROGRAM PELATIHAN") lalu
        # menimpanya dengan formula skor -- itulah sumber baris header yang
        # tiba-tiba terisi angka di sheet Penyelenggaraan.
        allowed_labels = SECTION_LABELS.get(sheet_name, list(row_map.keys()))
        for label in allowed_labels:
            if label not in row_map:
                continue
            master_row = row_map[label]
            question_text = canonical_questions.get(label)
            if not question_text:
                continue
            view_row = _find_question_row(ws, question_text)
            if not view_row:
                continue
            for col in used_cols:
                cell = ws[f"{col}{view_row}"]
                cell.value = f"='{MASTER_SHEET}'!{col}{master_row}"
                cell.number_format = hide_zero_format


def generate_official_report(
    template_path: str,
    df_filtered,
    score_col_map: dict,
    canonical_questions: dict,
    program_name: str,
    training_title: str = "",
    date_text: str = "",
    instructor_name: str = "",
) -> bytes:
    n = len(df_filtered)
    if n == 0:
        raise ValueError("Tidak ada responden pada data terfilter.")

    capped = min(n, MAX_RESPONDENTS)
    use_secondary = capped > DEFAULT_CAPACITY
    all_cols = PRIMARY_COLS + (SECONDARY_COLS if use_secondary else [])

    wb = openpyxl.load_workbook(template_path)
    ws_master = wb[MASTER_SHEET]

    _fix_secondary_header_numbers(ws_master)
    _sync_titles_and_separator_lines(wb)

    ws_master["A6"] = REPORT_TITLE
    if training_title:
        ws_master["A7"] = _format_training_title(training_title)
    if program_name:
        ws_master["A8"] = _format_program_title(program_name)
    if date_text:
        ws_master["A9"] = date_text
    if instructor_name:
        ws_master["C31"] = f"Nama Instruktur : {instructor_name}"
        if "INSTRUKTUR" in wb.sheetnames:
            wb["INSTRUKTUR"]["C12"] = f"INSTRUKTUR : {instructor_name}"

    # ---- Cocokkan tiap pertanyaan skor dengan baris di sheet master ----
    row_map = {}
    for label, question in canonical_questions.items():
        if label not in score_col_map:
            continue
        row = _find_question_row(ws_master, question)
        if row:
            row_map[label] = row

    # ---- Tulis nilai peserta (literal) di sheet master ----
    values_by_label = {}
    for label, row in row_map.items():
        col_name = score_col_map[label]
        values = (
            df_filtered[col_name]
            .dropna()
            .astype(float)
            .round()
            .astype(int)
            .tolist()[:capped]
        )
        values_by_label[label] = values

        for col in all_cols:
            ws_master[f"{col}{row}"] = None
        for i, val in enumerate(values):
            ws_master[f"{all_cols[i]}{row}"] = val

    # ---- Tulis ulang formula rujukan-silang di sheet turunan (perbaikan Bug 1 & 2) ----
    _rewrite_cross_sheet_formulas(wb, row_map, canonical_questions, all_cols)

    # ---- Style chart Histogram (skala sumbu menyesuaikan data aktual) ----
    section_avgs = []
    for labels in SECTION_LABELS.values():
        means = [sum(values_by_label[l]) / len(values_by_label[l]) for l in labels if l in values_by_label and values_by_label[l]]
        if means:
            section_avgs.append(sum(means) / len(means))
    _style_histogram_chart(wb, training_title or program_name, section_avgs)

    # ---- Perbaiki pembagi (divisor) & rentang SUM di semua sheet ----
    actual_n = max((len(v) for v in values_by_label.values()), default=capped)
    _update_divisor_everywhere(wb, actual_n, use_secondary)

    wb.calculation.fullCalcOnLoad = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), {
        "respondents_used": actual_n,
        "respondents_total_filtered": n,
        "capped": n > MAX_RESPONDENTS,
        "used_secondary_columns": use_secondary,
        "questions_matched": len(row_map),
        "questions_total": len(canonical_questions),
    }